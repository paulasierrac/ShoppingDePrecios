"""
================================================================================
Funciones utilitarias globales - Shopping de Precios Locatel
Autor: Paula Sierra — Net Applications
Propiedad de Colsubsidio
================================================================================

Contiene:
  - write_log         : Escribe logs en archivo de texto (equivale a WriteLog AA)
  - enviar_correo     : Envia correos desde tabla EnvioCorreos en BD (equivale a EnviarCorreo AA)
  - excel_a_csv       : Convierte hoja de Excel a CSV (equivale a ExcelToCsv AA)
  - csv_a_excel       : Convierte CSV a archivo Excel (equivale a CsvToExcel AA)
  - cargar_tabla_envio_correos : Carga tabla EnvioCorreos desde Excel a BD
  - obtener_config    : Lee config.json y lo retorna como diccionario
  - conectar_bd       : Retorna conexion pyodbc a SQL Server
"""

import os
import logging
import socket
import getpass
import smtplib
import shutil
import pyodbc
import pandas as pd
from datetime import datetime
from pathlib import Path
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders


# ---------------------------------------------------------------------------
# Configuracion global
# ---------------------------------------------------------------------------

def obtener_config() -> dict:
    """
    Obtiene credenciales de Azure Key Vault (secretos shared) y retorna el
    diccionario bootstrap con '_db', '_correo' y 'Scheme'.
    El resto de parametros del proceso se carga desde la BD en HU00.

    Requiere en .env: VAULT_URL, TENANT_ID, CLIENT_ID, CLIENT_SECRET, DB_SCHEMA.
    """
    import sys as _sys
    _root = Path(__file__).resolve().parent.parent
    if str(_root) not in _sys.path:
        _sys.path.insert(0, str(_root))

    from dotenv import load_dotenv
    load_dotenv(_root / ".env", override=False)

    from Config.Configuracion import CargarVault
    secretos = CargarVault(
        nombres=["Dev-DBName", "Dev-DBPassword", "Dev-DBServer", "Dev-DBUser"],
        strip_prefix="Dev",
    )

    schema = os.environ.get("DB_SCHEMA", "ShoppingDePrecios")

    return {
        "_db": {
            "driver":    "ODBC Driver 17 for SQL Server",
            "server":    secretos.get("DBSERVER",   ""),
            "database":  secretos.get("DBNAME",     ""),
            "usuario":   secretos.get("DBUSER",     ""),
            "contrasena":secretos.get("DBPASSWORD", ""),
        },
        "_correo": {
            "servidor_smtp": secretos.get("SMTP_HOST",     os.environ.get("SMTP_HOST",     "")),
            "puerto":        int(secretos.get("SMTP_PORT", os.environ.get("SMTP_PORT",     "587"))),
            "usar_tls":      True,
            "usuario":       secretos.get("SMTP_USER",     os.environ.get("SMTP_USER",     "")),
            "contrasena":    secretos.get("SMTP_PASSWORD", os.environ.get("SMTP_PASSWORD", "")),
        },
        "Scheme": f"[{schema}]",
    }


def conectar_bd_debug(config: dict) -> pyodbc.Connection:
    """En debug, conecta al SQL Server de desarrollo (credenciales Dev-* del vault)."""
    return conectar_bd(config)


def conectar_bd(config: dict) -> pyodbc.Connection:
    """Retorna conexion pyodbc a SQL Server usando los parametros del config."""
    db = config["_db"]
    drivers_disponibles = pyodbc.drivers()
    # Preferir driver moderno; caer al generico si no esta instalado
    driver = db.get("driver", "ODBC Driver 17 for SQL Server")
    if driver not in drivers_disponibles:
        for fallback in ("ODBC Driver 18 for SQL Server", "ODBC Driver 13 for SQL Server", "SQL Server"):
            if fallback in drivers_disponibles:
                driver = fallback
                break
    conn_str = (
        f"DRIVER={{{driver}}};"
        f"SERVER={db['server']};"
        f"DATABASE={db['database']};"
        f"UID={db['usuario']};"
        f"PWD={db['contrasena']}"
    )
    return pyodbc.connect(conn_str, autocommit=False)


# ---------------------------------------------------------------------------
# WriteLog
# ---------------------------------------------------------------------------

def write_log(in_state: str, in_message_log: str, in_task_name: str, in_config: dict) -> None:
    """
    Equivalente al bot WriteLog de AA.
    Siempre imprime a consola. Escribe a archivo si PathLog esta accesible.

    Formato de linea:
        dd/MM/yyyy HH:mm:ss | ESTADO | Mensaje | CodigoRobot | TaskName | Maquina

    Parametros:
        in_state       : "Info", "Error", "Warning", "Business"
        in_message_log : Texto del mensaje
        in_task_name   : Nombre de la tarea (equivale a $System:AATaskName$)
        in_config      : Diccionario de configuracion (ioConfig de AA)
    """
    activar = str(in_config.get("ActivarLog", "true")).lower()
    if activar != "true":
        return

    timestamp    = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
    maquina      = socket.gethostname()
    usuario      = getpass.getuser()
    codigo_robot = in_config.get("CodigoRobot", "")
    linea = (
        f"{timestamp} | {in_state} | {in_message_log} | "
        f"{codigo_robot} | {in_task_name} | {maquina}\n"
    )

    # Siempre a consola — independiente de PathLog
    print(linea.rstrip())

    # A archivo solo si PathLog esta configurado y accesible
    path_log = in_config.get("PathLog", "")
    if not path_log:
        return

    try:
        Path(path_log).mkdir(parents=True, exist_ok=True)
        fecha_archivo  = datetime.now().strftime("%Y%m%d")
        nombre_archivo = f"Log_{maquina}_{usuario}_{fecha_archivo}.txt"
        ruta_archivo   = os.path.join(path_log, nombre_archivo)
        with open(ruta_archivo, "a", encoding="ansi", errors="replace") as f:
            f.write(linea)
    except Exception as _log_err:
        print(f"[write_log] No se pudo escribir en archivo: {_log_err}")


# ---------------------------------------------------------------------------
# EnviarCorreo
# ---------------------------------------------------------------------------

def enviar_correo(
    in_config: dict,
    i_cod_email: int,
    i_from_address: str,
    i_replace_in_message: dict = None,
    i_replace_in_subject: dict = None,
    i_html_format: bool = False,
    i_attachment: list = None
) -> str:
    """
    Equivalente al bot EnviarCorreo de AA.

    Lee la fila de la tabla EnvioCorreos con CodEmailParameter = i_cod_email,
    reemplaza placeholders en asunto/cuerpo con i_replace_in_* y envia el correo.

    Retorna '' si exitoso, o mensaje de error si fallo.
    """
    out_system_exception = ""
    if i_replace_in_message is None:
        i_replace_in_message = {}
    if i_replace_in_subject is None:
        i_replace_in_subject = {}
    if i_attachment is None:
        i_attachment = []

    task_name = "EnviarCorreo"
    write_log("Info", f"Inicio envio correo cod={i_cod_email}", task_name, in_config)

    try:
        conn   = conectar_bd(in_config)
        cursor = conn.cursor()

        esquema = in_config.get("Esquema") or in_config.get("Scheme", "[ShoppingDePrecios]")
        tabla   = in_config.get("TablaEnvioCorreos", "EnvioCorreos")

        cursor.execute(
            f"SELECT * FROM {esquema}.{tabla} WHERE CodEmailParameter = ?",
            (str(i_cod_email),)
        )
        cols = [col[0] for col in cursor.description]
        row  = cursor.fetchone()

        if row is None:
            msg_err = (
                f"enviar_correo: tabla {esquema}.{tabla} no tiene "
                f"registro para CodEmailParameter={i_cod_email}. "
                f"Verificar que cargar_tabla_envio_correos se ejecuto correctamente."
            )
            write_log("Warning", msg_err, task_name, in_config)
            conn.close()
            return msg_err

        fila = dict(zip(cols, row))
        conn.close()

        subject = fila.get("AsuntoEmailParameter", "")
        body    = fila.get("BodyEmailParameter", "")
        to      = fila.get("TOEmailParameter", "")
        cc      = fila.get("CCEmailParameter", "")
        bcc     = fila.get("BCCEmailParameter", "")

        for key, val in i_replace_in_subject.items():
            subject = subject.replace(key, str(val))
        for key, val in i_replace_in_message.items():
            body = body.replace(key, str(val))

        mime_msg = MIMEMultipart()
        mime_msg["From"]    = i_from_address
        mime_msg["To"]      = to
        mime_msg["Subject"] = subject
        if cc:
            mime_msg["Cc"] = cc
        if bcc:
            mime_msg["Bcc"] = bcc

        mime_msg.attach(MIMEText(body, "html" if i_html_format else "plain", "utf-8"))

        for ruta_adj in i_attachment:
            if os.path.isfile(ruta_adj):
                with open(ruta_adj, "rb") as f:
                    parte = MIMEBase("application", "octet-stream")
                    parte.set_payload(f.read())
                encoders.encode_base64(parte)
                parte.add_header(
                    "Content-Disposition",
                    f'attachment; filename="{os.path.basename(ruta_adj)}"'
                )
                mime_msg.attach(parte)

        destinatarios = []
        for campo in [to, cc, bcc]:
            if campo:
                destinatarios.extend(
                    [d.strip() for d in campo.replace(";", ",").split(",") if d.strip()]
                )

        correo_cfg = in_config.get("_correo", {})
        servidor   = correo_cfg.get("servidor_smtp", "")
        puerto     = int(correo_cfg.get("puerto", 587))
        usar_tls   = correo_cfg.get("usar_tls", True)
        usr_smtp   = correo_cfg.get("usuario", i_from_address)
        pwd_smtp   = correo_cfg.get("contrasena", "")

        if not servidor:
            msg_err = (
                "enviar_correo: SMTP no configurado. "
                "Agregar SMTP_HOST, SMTP_PORT, SMTP_USER, SMTP_PASSWORD en .env "
                "o en Azure Key Vault."
            )
            write_log("Error", msg_err, task_name, in_config)
            return msg_err

        write_log("Info",
                  f"Conectando a SMTP {servidor}:{puerto} como {usr_smtp}",
                  task_name, in_config)

        with smtplib.SMTP(servidor, puerto, timeout=30) as server:
            server.ehlo()
            if usar_tls:
                server.starttls()
                server.ehlo()
            if usr_smtp and pwd_smtp:
                server.login(usr_smtp, pwd_smtp)
            server.sendmail(i_from_address, destinatarios, mime_msg.as_string())

        write_log("Info",
                  f"Correo enviado correctamente cod={i_cod_email} -> {destinatarios}",
                  task_name, in_config)

    except Exception as e:
        out_system_exception = str(e)
        write_log("Error", f"enviar_correo cod={i_cod_email}: {e}", task_name, in_config)

    return out_system_exception


# ---------------------------------------------------------------------------
# ExcelToCsv
# ---------------------------------------------------------------------------

def excel_a_csv(
    in_path_excel: str,
    in_path_out_csv: str,
    in_name_sheet: str,
    in_last_column: str = None,
    in_decimal_separator: str = ",",
    in_thousands_separator: str = ".",
    in_config: dict = None
) -> str:
    """
    Equivalente al bot ExcelToCsv de AA.
    Lee una hoja de Excel y la guarda como CSV con separador ';'.

    Retorna '' si exitoso, mensaje de error si fallo.
    """
    task_name = "ExcelToCsv"
    if in_config is None:
        in_config = {}

    try:
        write_log("Info", "INICIO ExcelToCsv", task_name, in_config)

        # Leer la hoja indicada
        df = pd.read_excel(
            in_path_excel,
            sheet_name=in_name_sheet,
            dtype=str,
            header=0
        )
        df = df.fillna("")

        # Si se indica columna limite, recortar columnas
        if in_last_column:
            col_idx = _letra_a_indice(in_last_column)
            if col_idx < len(df.columns):
                df = df.iloc[:, :col_idx]

        # Guardar CSV con separador punto y coma, codificacion ANSI (cp1252)
        df.to_csv(
            in_path_out_csv,
            sep=";",
            index=False,
            encoding="cp1252",
            errors="replace"
        )
        write_log("Info", "FIN ExcelToCsv", task_name, in_config)
        return ""

    except Exception as e:
        msg = f"Error en ExcelToCsv: {e}"
        write_log("Error", msg, task_name, in_config)
        return msg


def _letra_a_indice(letra: str) -> int:
    """Convierte letra de columna Excel (A, B, ..., Z, AA, ...) a indice entero base-0."""
    letra = letra.upper().strip()
    resultado = 0
    for c in letra:
        resultado = resultado * 26 + (ord(c) - ord('A') + 1)
    return resultado


# ---------------------------------------------------------------------------
# CsvToExcel
# ---------------------------------------------------------------------------

def csv_a_excel(
    in_path_temporal_csv: str,
    in_path_excel: str,
    in_extension_excel: str = "xlsx",
    in_new_name_sheet: str = "Hoja1",
    in_create_new_excel: bool = True,
    in_number_sheet: int = 1,
    in_type_of_copy: bool = False,
    in_config: dict = None
) -> str:
    """
    Equivalente al bot CsvToExcel de AA.
    Convierte un CSV a archivo Excel (.xlsx).

    Retorna '' si exitoso, mensaje de error si fallo.
    """
    task_name = "CsvToExcel"
    if in_config is None:
        in_config = {}

    try:
        write_log("Info", "INICIO CsvToExcel", task_name, in_config)

        df = pd.read_csv(in_path_temporal_csv, sep=";", dtype=str, encoding="cp1252", encoding_errors="replace")
        df = df.fillna("")

        if in_create_new_excel:
            # Crear nuevo archivo Excel
            with pd.ExcelWriter(in_path_excel, engine="openpyxl") as writer:
                df.to_excel(writer, sheet_name=in_new_name_sheet, index=False)
        else:
            # Agregar hoja a Excel existente
            from openpyxl import load_workbook
            if os.path.isfile(in_path_excel):
                wb = load_workbook(in_path_excel)
            else:
                from openpyxl import Workbook
                wb = Workbook()
                # Eliminar la hoja por defecto si existe
                if "Sheet" in wb.sheetnames:
                    del wb["Sheet"]

            if in_new_name_sheet in wb.sheetnames:
                del wb[in_new_name_sheet]

            ws = wb.create_sheet(title=in_new_name_sheet)
            # Escribir encabezados
            ws.append(list(df.columns))
            for _, fila in df.iterrows():
                ws.append(list(fila))
            wb.save(in_path_excel)

        write_log("Info", "FIN CsvToExcel", task_name, in_config)
        return ""

    except Exception as e:
        msg = f"Error en CsvToExcel: {e}"
        write_log("Error", msg, task_name, in_config)
        return msg


# ---------------------------------------------------------------------------
# CargarTablaEnvioCorreos
# ---------------------------------------------------------------------------

def cargar_tabla_envio_correos(in_config: dict) -> str:
    """
    Equivalente al bot CargarTablaEnvioCorreos de AA.

    Lee el archivo Excel de EnvioCorreos, lo convierte a CSV temporal y
    hace UPSERT en la tabla EnvioCorreos de la BD.

    Retorna '' si exitoso, mensaje de error si fallo.
    """
    task_name = "CargarTablaEnvioCorreos"
    out_system_exception = ""

    write_log("Info", "Inicia funcion CargarTablaEnvioCorreos", task_name, in_config)

    try:
        # Ruta del archivo de insumo de correos — usa rutas completas desde BD
        archivo_envio_correos = in_config.get("ArchivoEnvioCorreos", "EnvioCorreos.xlsx")
        ruta_insumo = os.path.join(in_config.get("RutaInsumos", ""), archivo_envio_correos)

        if not os.path.isfile(ruta_insumo):
            write_log("Info", f"NO existe el archivo de la ruta ({ruta_insumo})", task_name, in_config)
            write_log("Info", "Finaliza funcion CargarTablaEnvioCorreos (sin archivo)", task_name, in_config)
            return ""  # No es error critico - puede que ya este cargada

        write_log("Info", f"Existe el archivo de la ruta ({ruta_insumo})", task_name, in_config)

        # Convertir Excel a CSV temporal
        ruta_csv = os.path.join(in_config.get("RutaTemp", ""), "InsumoT.csv")

        if os.path.isfile(ruta_csv):
            os.remove(ruta_csv)

        sheet_envio = in_config.get("SheetEnvioCorreos", "EnvioCorreos")
        err = excel_a_csv(
            in_path_excel=ruta_insumo,
            in_path_out_csv=ruta_csv,
            in_name_sheet=sheet_envio,
            in_last_column="H",
            in_thousands_separator=".",
            in_config=in_config
        )
        if err:
            raise Exception(f"Error convirtiendo Excel a CSV: {err}")

        # Cargar datos a BD
        df = pd.read_csv(ruta_csv, sep=";", dtype=str, encoding="cp1252", encoding_errors="replace")
        df = df.fillna("")
        # Normalizar nombres de columna (soporte para Excel en espanol e ingles)
        col_map = {
            "IdCorreo":        "CodEmailParameter",
            "CorreoPara":      "TOEmailParameter",
            "CorreoParaCopia": "CCEmailParameter",
            "CorreoParOculto": "BCCEmailParameter",
            "Asunto":          "AsuntoEmailParameter",
            "Contenido":       "BodyEmailParameter",
            "BooleanHTML":     "IsHTMLEmailParameter",
        }
        df = df.rename(columns={k: v for k, v in col_map.items() if k in df.columns})
        # Eliminar filas con CodEmailParameter vacio o nulo
        if "CodEmailParameter" in df.columns:
            df = df[df["CodEmailParameter"].str.strip() != ""]

        conn = conectar_bd(in_config)
        cursor = conn.cursor()

        esquema = in_config.get("Esquema") or in_config.get("Scheme", "[ShoppingDePrecios]")
        tabla = in_config.get("TablaEnvioCorreos", "EnvioCorreos")

        # Crear tabla temporal
        cursor.execute("""
            IF OBJECT_ID('tempdb.dbo.#Temporal', 'U') IS NOT NULL
                DROP TABLE #Temporal;
            CREATE TABLE #Temporal(
                [CodEmailParameter] [varchar](10) NULL,
                [TOEmailParameter] [varchar](max) NULL,
                [CCEmailParameter] [varchar](max) NULL,
                [BCCEmailParameter] [varchar](max) NULL,
                [AsuntoEmailParameter] [varchar](250) NULL,
                [BodyEmailParameter] [varchar](max) NULL,
                [IsHTMLEmailParameter] [varchar](10) NULL
            )
        """)

        cols_esperadas = [
            "CodEmailParameter", "TOEmailParameter", "CCEmailParameter",
            "BCCEmailParameter", "AsuntoEmailParameter", "BodyEmailParameter",
            "IsHTMLEmailParameter"
        ]

        for _, row in df.iterrows():
            vals = [row.get(c, "") for c in cols_esperadas]
            placeholders = ",".join(["?" for _ in vals])
            cursor.execute(f"INSERT INTO #Temporal VALUES ({placeholders})", vals)

        # Depurar filas invalidas
        cursor.execute("""
            DELETE FROM #Temporal
            WHERE CodEmailParameter LIKE '' OR CodEmailParameter IS NULL
        """)

        # INSERT para nuevos codigos
        cursor.execute(f"""
            INSERT INTO {esquema}.{tabla}
                ([IdEmailParameter],[CodEmailParameter],[TOEmailParameter],[CCEmailParameter],
                 [BCCEmailParameter],[AsuntoEmailParameter],[BodyEmailParameter],[IsHTMLEmailParameter])
            SELECT TRY_CAST(t.[CodEmailParameter] AS INT),
                   t.[CodEmailParameter],t.[TOEmailParameter],t.[CCEmailParameter],
                   t.[BCCEmailParameter],t.[AsuntoEmailParameter],t.[BodyEmailParameter],
                   TRY_CAST(t.[IsHTMLEmailParameter] AS BIT)
            FROM #Temporal t
            WHERE NOT EXISTS (
                SELECT 1 FROM {esquema}.{tabla} c WHERE c.CodEmailParameter = t.CodEmailParameter
            )
        """)

        # UPDATE para codigos existentes
        cursor.execute(f"""
            UPDATE {esquema}.{tabla}
            SET [CodEmailParameter]=t2.[CodEmailParameter],
                [TOEmailParameter]=t2.[TOEmailParameter],
                [CCEmailParameter]=t2.[CCEmailParameter],
                [BCCEmailParameter]=t2.[BCCEmailParameter],
                [AsuntoEmailParameter]=t2.[AsuntoEmailParameter],
                [BodyEmailParameter]=t2.[BodyEmailParameter],
                [IsHTMLEmailParameter]=t2.[IsHTMLEmailParameter]
            FROM {esquema}.{tabla} t1
            INNER JOIN #Temporal t2 ON t1.CodEmailParameter = t2.CodEmailParameter
        """)

        conn.commit()
        conn.close()

        write_log("Info", "Finaliza funcion CargarTablaEnvioCorreos", task_name, in_config)
        out_system_exception = ""

    except Exception as e:
        out_system_exception = str(e)
        write_log("Error", f"Error en CargarTablaEnvioCorreos: {e}", task_name, in_config)
        write_log("Info", "Finaliza funcion CargarTablaEnvioCorreos", task_name, in_config)

    return out_system_exception
